#!/usr/bin/env python3
"""Generate a timetable Excel file.

Usage:
    python generate_timetable.py        # uses internal sample data
    python generate_timetable.py data.csv  # reads CSV with same columns
"""

import sys
import pandas as pd


def default_entries():
    """Return the hard-coded schedule from the user request."""
    # columns: Time, Content, Speaker
    return [
        {"Time": "09:00~09:30", "Content": "報到", "Speaker": ""},
        {"Time": "09:30~09:40", "Content": "開場致詞", "Speaker": "王教授"},
        {"Time": "09:40~10:05", "Content": "邁向6G 的AI-RAN及O-RAN 趨勢介紹", "Speaker": "劉教授"},
        {"Time": "10:05~10:30", "Content": "下世代B5G/6G專網應用與未來趨勢", "Speaker": "陳教授"},
        {"Time": "10:30~10:50", "Content": "Break", "Speaker": ""},
        {"Time": "10:50~11:20", "Content": "從O-RAN到AI-RAN 智慧通訊的節能應用", "Speaker": "教學團隊"},
        {"Time": "11:20~12:00", "Content": "O-RAN環境和各模組化功能介紹", "Speaker": "教學團隊"},
        {"Time": "12:00~13:30", "Content": "Lunch", "Speaker": ""},
        {"Time": "13:30~14:00", "Content": "O-RAN 的市場應用案例", "Speaker": "教學團隊"},
        {"Time": "14:00~14:30", "Content": "O-RAN OSC環境建置教學", "Speaker": "教學團隊"},
        {"Time": "14:30~14:50", "Content": "Break", "Speaker": ""},
        {"Time": "14:50~15:50", "Content": "O-RAN OSC第三方應用程式 xApps建置教學", "Speaker": "教學團隊"},
        {"Time": "15:50~16:30", "Content": "現場討論時間", "Speaker": "教學團隊"},
    ]


def generate_timetable(entries, output_path="timetable.xlsx"):
    """Write the entries list of dicts to an Excel file."""
    df = pd.DataFrame(entries)

    # cells may contain arbitrary text; no modifications performed

    # write with formatting via openpyxl
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Timetable")
        workbook = writer.book
        sheet = writer.sheets["Timetable"]

        # set column widths so that the longest line in each column fits
        for idx, col in enumerate(df.columns, 1):
            maxlen = len(col)
            for val in df[col].astype(str):
                # consider each explicit line separately
                for part in val.split("\n"):
                    maxlen = max(maxlen, len(part))
            width = maxlen + 5
            sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

        # apply formatting to all cells
        from openpyxl.styles import Alignment, Border, Side
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # track desired heights per row, accounting for wrap
        from math import ceil
        row_heights = {}
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = align
                cell.border = border
                # calculate lines needed for this cell
                text = str(cell.value or "")
                # count explicit breaks
                lines = text.count("\n") + 1
                # estimate wrap due to width
                col_letter = cell.column_letter
                col_width = sheet.column_dimensions[col_letter].width or 10
                # assume roughly one character per unit width
                lines = max(lines, ceil(len(text) / col_width))
                row_idx = cell.row
                row_heights[row_idx] = max(row_heights.get(row_idx, 0), lines * 15)
        # apply computed row heights
        for idx, height in row_heights.items():
            sheet.row_dimensions[idx].height = height

        # merge only the content and speaker columns for specified rows
        merge_for = {"報到", "Break", "Lunch"}
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            if row.Content in merge_for:
                # keep Time column separate; merge columns B and C instead
                sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)

    print(f"Timetable written to {output_path}")


def read_csv(path):
    """Read a CSV file and return list of dicts.

    The CSV should have columns named Time, Content, Speaker (case-insensitive).
    """
    df = pd.read_csv(path)
    # ensure columns
    expected = ["Time", "Content", "Speaker"]
    for col in expected:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")
    return df.to_dict(orient="records")


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
        try:
            entries = read_csv(path)
        except Exception as e:
            print(f"Failed to read CSV: {e}")
            sys.exit(1)
    else:
        entries = default_entries()

    generate_timetable(entries)


if __name__ == "__main__":
    main()
